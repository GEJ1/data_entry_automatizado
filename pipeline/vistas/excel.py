"""
Vistas en Excel para revision humana. Se generan DESDE el JSONL.

Ojo con la direccion de la flecha: el JSONL es la fuente de verdad y el Excel
es una foto derivada. Las correcciones NO vuelven editando el Excel: se corrige
la regla en el codigo y se re-corre. Si el Excel fuera editable de vuelta, a la
segunda corrida nadie sabria cual de los dos tiene razon.

Cuatro hojas, cuatro preguntas distintas:
  * Resumen   -> "¿esta todo el mundo?"   (una fila por cliente, para reconciliar)
  * Detalle   -> "¿este dato quedo bien?" (una fila por referencia)
  * Alertas   -> "¿y los antecedentes?"   (una fila por alerta)
  * Problemas -> "¿que tengo que mirar?"  (solo lo dudoso, idealmente vacia)

Detalle y Alertas son las dos cosas que efectivamente se cargan en la web, una
por formulario. Van en hojas separadas porque son entidades distintas y se
revisan distinto, no por prolijidad.

La hoja Problemas es la mas importante: si esta vacia, se puede cargar tranquilo.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pipeline.dominio.esquema import Cliente

RELLENO_ENCABEZADO = PatternFill("solid", fgColor="1F4E79")
FUENTE_ENCABEZADO = Font(color="FFFFFF", bold=True)
RELLENO_ALERTA = PatternFill("solid", fgColor="FCE4E4")

COLS_RESUMEN = ["CUIT", "CUIT ok", "Cliente", "Emisión", "Referencias",
                "Descartadas (inactivas)", "Alertas", "Problemas", "Origen"]

COLS_DETALLE = ["CUIT", "Cliente", "Solicitud", "Fecha", "Informante",
                "¿Es cliente?", "CO ($)", "CT ($)", "CO (USD)", "CT (USD)",
                "Condición de venta", "Plazo", "Concepto", "Antigüedad",
                "Problemas"]

COLS_ALERTAS = ["CUIT", "Cliente", "Fecha", "Alertante", "Tipo", "Estado",
                "Monto", "Comentarios"]

COLS_PROBLEMAS = ["CUIT", "Cliente", "Problema", "Origen"]


def _encabezar(hoja, columnas: list[str]) -> None:
    hoja.append(columnas)
    for celda in hoja[1]:
        celda.fill = RELLENO_ENCABEZADO
        celda.font = FUENTE_ENCABEZADO
        celda.alignment = Alignment(vertical="center", wrap_text=True)
    hoja.freeze_panes = "A2"                      # el encabezado no se va al scrollear
    hoja.auto_filter.ref = hoja.dimensions        # filtros, que es como se revisa


def _ajustar_anchos(hoja, maximo: int = 42) -> None:
    for i, columna in enumerate(hoja.columns, start=1):
        largo = max((len(str(c.value)) for c in columna if c.value is not None),
                    default=8)
        hoja.column_dimensions[get_column_letter(i)].width = min(largo + 2, maximo)


def _texto(valor) -> str:
    """None -> '' . Excel muestra 'None' si no, y eso se lee como un dato."""
    return "" if valor is None else str(valor)


def generar(clientes: list[Cliente], salida: str | Path) -> Path:
    """Escribe el .xlsx con las tres hojas. Devuelve la ruta."""
    libro = Workbook()

    # -- Resumen: una fila por cliente ---------------------------------------
    resumen = libro.active
    resumen.title = "Resumen"
    _encabezar(resumen, COLS_RESUMEN)
    for c in clientes:
        resumen.append([c.cuit, "sí" if c.cuit_ok else "NO", c.nombre,
                        c.emision.strftime("%d/%m/%Y"), len(c.referencias),
                        c.descartadas, len(c.alertas), len(c.problemas), c.origen])
        if c.problemas:
            for celda in resumen[resumen.max_row]:
                celda.fill = RELLENO_ALERTA

    # -- Detalle: una fila por referencia (lo que efectivamente se carga) -----
    detalle = libro.create_sheet("Detalle")
    _encabezar(detalle, COLS_DETALLE)
    for c in clientes:
        for r in c.referencias:
            detalle.append([
                c.cuit, c.nombre, r.solicitud,
                _texto(r.fecha and r.fecha.strftime("%d/%m/%Y")), r.informante,
                "sí" if r.es_cliente else "no",
                _texto(r.co_ars), _texto(r.ct_ars),
                _texto(r.co_usd), _texto(r.ct_usd),
                r.condicion_venta, _texto(r.plazo), r.concepto,
                _texto(r.antiguedad and r.antiguedad.strftime("%d/%m/%Y")),
                "; ".join(r.problemas),
            ])
            if r.problemas:
                for celda in detalle[detalle.max_row]:
                    celda.fill = RELLENO_ALERTA

    # -- Alertas: una fila por alerta (el otro formulario que se carga) ------
    alertas = libro.create_sheet("Alertas")
    _encabezar(alertas, COLS_ALERTAS)
    for c in clientes:
        for a in c.alertas:
            alertas.append([
                c.cuit, c.nombre,
                _texto(a.fecha and a.fecha.strftime("%d/%m/%Y")), a.alertante,
                a.tipo, a.estado, _texto(a.monto), a.comentarios,
            ])

    # -- Problemas: solo lo dudoso -------------------------------------------
    problemas = libro.create_sheet("Problemas")
    _encabezar(problemas, COLS_PROBLEMAS)
    for c in clientes:
        for p in c.problemas:
            problemas.append([c.cuit, c.nombre, p, c.origen])

    for hoja in (resumen, detalle, alertas, problemas):
        _ajustar_anchos(hoja)

    salida = Path(salida)
    salida.parent.mkdir(parents=True, exist_ok=True)
    libro.save(salida)
    return salida
